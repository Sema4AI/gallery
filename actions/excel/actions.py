"""Set of actions interacting with local Excel files.

These are currently capable of managing workbooks and their worksheets, while
retrieving and setting data from/into their cells.
"""

import contextlib
from copy import copy
import re
from pathlib import Path

from robocorp import excel, log
from robocorp.excel.tables import Table as ExcelTable
from sema4ai.actions import ActionError, Response, action
from openpyxl import load_workbook
from datetime import datetime
import openpyxl.utils
from models import (
    Row,
    Schema,
    Sheet,
    Table,
    Header,
    CrossReferenceResult,
    UserAndTime,
)
from robocorp.excel.worksheet import Worksheet
from sema4ai.crossplatform import trigger_excel_save_on_app


def _search_excel_file(file_path: str, is_expected: bool = True) -> Path:
    # Guesses the local Excel file path if the provided one isn't specific.
    path = Path(file_path)

    # Search for an existing Excel when the extension isn't provided, as usually the
    #  LLM may not receive a full path with directories, name and extension specified.
    #  (e.g.: "Create a new Excel file called 'data' please.")
    if not path.suffix:
        file_name = path.stem
        search_dir = path.parent
        try:
            path = next(search_dir.glob(f"{file_name}.xls*"))
        except StopIteration:
            if is_expected:
                # Since the file is expected to exist, we let the caller know that it
                #  wasn't found.
                raise FileNotFoundError(
                    f"no Excel file matching {file_name!r} found in: {search_dir}"
                )
            else:
                # Otherwise we just pick a default favoured extension.
                path = Path(f"{file_path}.xlsx")

    return path.expanduser().resolve()


@contextlib.contextmanager
def _capture_error(exc_type: type = Exception):
    # Captures given error types and re-raises them as action errors in order to return
    #  a result containing an error message through the Action Server.
    try:
        yield
    except exc_type as exc:
        log.exception(exc)
        raise ActionError(exc) from exc


@contextlib.contextmanager
def _open_workbook(
    file_path: str,
    sheet_name: str | None = None,
    save: bool = True,
    sheet_required: bool = False,
):
    # Opens a workbook, provides the control to a sheet optionally, then saves the
    #  changes that may have occurred in the caller.
    print(f"Opening workbook with file path: {file_path}")
    with _capture_error(FileNotFoundError):
        path = _search_excel_file(file_path)
        workbook = excel.open_workbook(path)

    if sheet_name:
        print(f"Opening worksheet with name: {sheet_name}")
        with _capture_error(KeyError):
            worksheet = workbook.worksheet(sheet_name)
        yield workbook, worksheet
    elif sheet_required:
        # Ensures the provided name is valid and has content.
        raise ActionError("a sheet name has to be specified")
    else:
        yield workbook, None

    if save:
        print(f"Saving workbook changes in path: {path}")
        workbook.save(path, overwrite=True)


def _column_to_number(column) -> int:
    """Convert Excel-style column letter to number (e.g., 'A' -> 1, 'AD' -> 30)."""
    number = 0
    for char in column:
        number = number * 26 + (ord(char) - ord("A") + 1)
    return number


def _extract_column_and_row(cell):
    """Extract the column letters and row number from a cell reference like 'AD2'."""
    match = re.match(r"([A-Z]+)(\d+)", cell, re.I)
    if match:
        return match.groups()
    else:
        raise ActionError("Invalid cell reference")


@action(is_consequential=True)
def create_workbook(file_path: str, sheet_name: str = "") -> Response[str]:
    """Create a new local Excel file defined as workbook.

    This will overwrite an already existing file and will try to guess the extension
    (xls or xlsx) if is not specified. It also creates a sheet with a default name
    which can be customized by providing an explicit `sheet_name`.

    Args:
        file_path: The file name or a local path pointing to the workbook file.
        sheet_name: The name of the initial sheet in the workbook. Leave it blank for
            relying on the default name ("Sheet").

    Returns:
        The absolute path of the newly created local Excel file.
    """
    path = _search_excel_file(file_path, is_expected=False)
    extension = path.suffix.strip(".").lower()
    workbook = excel.create_workbook(
        "xls" if extension == "xls" else "xlsx", sheet_name=sheet_name or None
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return Response(result=f"Created new Excel file in location: {path}")


@action(is_consequential=True)
def delete_workbook(file_path: str) -> Response[str]:
    """Delete an existing local Excel file defined as workbook.

    Will return an error message if the passed file is not found.

    Args:
        file_path: The file name or a local path pointing to the workbook file.

    Returns:
        The absolute path of the removed local Excel file.
    """
    print(f"Removing Excel file in path: {file_path}")
    with _capture_error(FileNotFoundError):
        path = _search_excel_file(file_path)
        path.unlink()

    return Response(result=f"Removed Excel file at location: {path}")


@action(is_consequential=True)
def create_worksheet(file_path: str, sheet_name: str) -> Response[str]:
    """Create a new sheet in an already existing workbook.

    Will return an error message if the sheet already exists.

    Args:
        file_path: The file name or a local path pointing to the workbook file.
        sheet_name: The name of the new sheet to be added in the workbook.

    Returns:
        The name of the newly created worksheet.
    """
    with _open_workbook(file_path) as (workbook, _):
        print(f"Creating sheet with name: {sheet_name}")
        with _capture_error(ValueError):
            workbook.create_worksheet(sheet_name)

    return Response(result=f"Successfully created sheet {sheet_name!r}!")


@action(is_consequential=True)
def delete_worksheet(file_path: str, sheet_name: str) -> Response[str]:
    """Remove a sheet from an already existing workbook.

    Will return an error message if the sheet is not present.

    Args:
        file_path: The file name or a local path pointing to the workbook file.
        sheet_name: The name of the sheet you want removed from the workbook.

    Returns:
        The name of the just removed worksheet.
    """
    with _open_workbook(file_path) as (workbook, _):
        print(f"Removing sheet with name: {sheet_name}")
        with _capture_error(KeyError):
            workbook.remove_worksheet(sheet_name)

    return Response(result=f"Successfully deleted sheet {sheet_name!r}!")


@action(is_consequential=True)
def add_rows(file_path: str, sheet_name: str, data_table: Table) -> Response[str]:
    """Add rows in an already existing worksheet of a workbook.

    This will append the rows at the next empty index available. Only use this method on an empty file or
    when you want to add data to the end of the file. If you want to add a new column or manipulate data
    use `update_sheet_rows` instead.

    Args:
        file_path: The file name or a local path pointing to the workbook file.
        sheet_name: The name of the sheet you want to store the rows into.
        data_table: A 2D matrix containing the sheet cell values.

    Returns:
        How many rows were added in the sheet.
    """
    table = ExcelTable(data=data_table.as_list())

    effect = f"{len(table)} row(s) to sheet {sheet_name!r}"
    print(f"Adding {effect}...")

    with _open_workbook(file_path, sheet_name, sheet_required=True) as (_, worksheet):
        worksheet.append_rows_to_worksheet(table)

    return Response(result=f"{effect} were added successfully!")


@action(is_consequential=True)
def update_sheet_rows(
    file_path: str,
    sheet_name: str,
    start_cell: str,
    data: Table,
    overwrite: bool = False,
) -> Response[str]:
    """Update a cell or a range of cells in a worksheet using A1 notation.

    Args:
        file_path: The file name or a local path pointing to the workbook file.
        sheet_name: The name of the sheet you want to store the rows into.
        start_cell: The cell from where to start the update. The end will be determined based on each row's length.
        data: Data to be inserted into the cell or cells.
        overwrite: If True, the data will overwrite the existing data in the cells.

    Example:
        ```python
            update_sheet_rows("Orders", "January Sheet", "A1", [["Order number"]])
            update_sheet_rows("Orders", "January Sheet", "B1", [["Item"]])
            update_sheet_rows("Orders", "January Sheet", "A2", [[5, "Macbook Pro"], [6, "iPhone"]])
        ```

    Returns:
         Message indicating the success or failure of the operation.
    """
    start_column, start_row = _extract_column_and_row(start_cell)
    start_column_number = _column_to_number(start_column)
    start_row = int(start_row)

    data = data if isinstance(data, list) else data.as_list()
    with _open_workbook(file_path, sheet_name, sheet_required=True) as (
        workbook,
        worksheet,
    ):
        for i, row_values in enumerate(data):
            current_row = start_row + i  # Move down rows

            for j, value in enumerate(row_values):
                column_number = start_column_number + j  # Move across columns
                start_column_letter = openpyxl.utils.get_column_letter(column_number)
                update_cell_value(
                    worksheet,
                    f"{start_column_letter}{current_row}",
                    str(value),
                    overwrite,
                )
        workbook.save(file_path)
    return Response(result="Rows were successfully updated.")


def update_cell_value(
    ws: Worksheet, cell_reference: str, value: str, overwrite: bool = False
) -> None:
    """Updates cell value while preserving all formatting

    If cell contains a formula, do not set overwrite to True unless user
    specifies that they want to overwrite the formula.
    Args:
        ws: Worksheet object
        cell_reference: Cell reference (e.g., 'A1')
        value: Value to insert into the cell
        overwrite: If True, the data will overwrite the existing formula in the cells.
    """
    worksheet = ws._workbook.excel.book.active
    cell = worksheet[cell_reference]

    # Check if the cell contains a formula
    has_existing_formula = isinstance(cell.value, str) and cell.value.startswith("=")
    if has_existing_formula and not overwrite:
        print(f"Cannot overwrite formula in {cell_reference}")
        return
    # Store all formatting attributes
    font = copy(cell.font)
    fill = copy(cell.fill)
    border = copy(cell.border)
    alignment = copy(cell.alignment)
    number_format = cell.number_format

    # Store comment if exists
    comment = cell.comment
    # Convert value to appropriate type based on current cell format
    if isinstance(value, str):
        if value.startswith("="):  # Formula
            converted_value = value
        elif number_format.endswith("%"):  # Percentage
            try:
                converted_value = float(value.rstrip("%")) / 100
            except ValueError:
                converted_value = value
        elif "General" in number_format or "0" in number_format:  # Number
            try:
                if "." in value:
                    converted_value = float(value)
                else:
                    converted_value = int(value)
            except ValueError:
                converted_value = value
        else:  # Default to string
            converted_value = value
    else:
        converted_value = value

    # Set the new value
    cell.value = converted_value

    # Reapply all formatting
    cell.font = font
    cell.fill = fill
    cell.border = border
    cell.alignment = alignment
    cell.number_format = number_format

    # Reapply comment if it existed
    if comment:
        cell.comment = comment


@action(is_consequential=False)
def get_cell(
    file_path: str, sheet_name: str, row: str, column: str
) -> Response[str | None]:
    """Retrieve the value of a cell from an already existing worksheet of a workbook.

    This action works similarly to `set_cell`, but instead of setting a value it
    actually returns the value found at the given position in the sheet.

    Args:
        file_path: The file name or a local path pointing to the workbook file.
        sheet_name: The name of the sheet you want to store the rows into.
        row: The row number starting with 1.
        column: The column letter or index.

    Returns:
        The value found at the given `row` and `column`.
    """
    with _open_workbook(file_path, sheet_name, save=False, sheet_required=True) as (
        _,
        worksheet,
    ):
        value = worksheet.get_cell_value(row, column)

    return Response(result=str(value))


@action(is_consequential=False)
def get_sheet_content(file_path: str, sheet_name: str) -> Response[Table]:
    """Retrieve the whole content of an already existing worksheet of a workbook.

    This action retrieves all the cells found in the sheet and returns them in a table
    structure.

    Args:
        file_path: The file name or a local path pointing to the workbook file.
        sheet_name: The name of the sheet you want to store the rows into.

    Returns:
        A table structure capturing the entire content from the requested sheet.
    """
    with _open_workbook(file_path, sheet_name, save=False, sheet_required=True) as (
        _,
        worksheet,
    ):
        sheet_table = worksheet.as_table()

    rows = []
    for sheet_row in sheet_table.iter_lists(with_index=False):
        rows.append(Row(cells=sheet_row))

    return Response(result=Table(rows=rows))


@action(is_consequential=False)
def get_workbook_schema(file_path: str) -> Response[Schema]:
    """Retrieve the workbook overview defined as schema.

    This action identifies the sheets found in the given workbook and for each one of
    them it presents the first non-empty row which usually represents the header of the sheet.

    Args:
        file_path: The file name or a local path pointing to the workbook file.

    Returns:
        A list of sheets where for every sheet its name and first non-empty row are given.
    """
    sheets = []
    path = _search_excel_file(file_path)

    # Use openpyxl directly for better performance

    try:
        workbook = load_workbook(filename=path, read_only=True)
        created = UserAndTime(
            user=(
                workbook.properties.creator.encode("utf-8")
                if workbook.properties.creator
                else ""
            ),
            time=workbook.properties.created,
        )
        last_modified = UserAndTime(
            user=(
                workbook.properties.last_modified_by.encode("utf-8")
                if workbook.properties.last_modified_by
                else ""
            ),
            time=workbook.properties.modified,
        )
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            min_col_letter = openpyxl.utils.get_column_letter(worksheet.min_column)
            max_col_letter = openpyxl.utils.get_column_letter(worksheet.max_column)
            data_range = f"{min_col_letter}{worksheet.min_row}:{max_col_letter}{worksheet.max_row}"
            sheets.append(
                Sheet(
                    name=sheet_name,
                    data_range=data_range,
                )
            )
    finally:
        workbook.close()

    return Response(
        result=Schema(
            sheets=sheets,
            created=created,
            last_modified=last_modified,
        )
    )


@action
def find_cross_reference(
    file_path: str, sheet_name: str, header1: Header, header2: Header
) -> CrossReferenceResult:
    """Find the cell references where header1 and header2 intersect in the same row.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to search
        header1: First header value(s) to find (can be a string or list of strings)
        header2: Second header value(s) to find (can be a string or list of strings)
    Returns:
        List of dictionaries containing the intersection descriptions and cell references
    """
    if isinstance(header1.value, list) and isinstance(header2.value, list):
        raise ValueError("Only one of header1 or header2 can be a list, not both.")

    # Ensure headers are lists for uniform processing
    if isinstance(header1.value, list):
        h1_list = header1.value
    else:
        h1_list = [header1.value]

    if isinstance(header2.value, list):
        h2_list = header2.value
    else:
        h2_list = [header2.value]

    # Dictionary to store all occurrences of each header
    header_positions = {h: {"rows": [], "cols": []} for h in h1_list + h2_list}

    # Load both normal and data_only workbooks
    wb = load_workbook(file_path, data_only=False)
    trigger_excel_save_on_app(file_path)
    wb_data = load_workbook(file_path, data_only=True)

    ws = wb[sheet_name]
    ws_data = wb_data[sheet_name]

    # Search for all headers in the worksheet
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            # Get both formula and calculated value
            cell_data = ws_data.cell(row=row_idx, column=col_idx)
            calculated_value = cell_data.value

            # Get the formatted value as it appears in Excel
            formatted_value = None
            if calculated_value is not None:
                if cell_data.is_date:
                    try:
                        format_str = _excel_format_to_strftime(cell.number_format)
                        formatted_value = calculated_value.strftime(format_str)
                    except (AttributeError, ValueError):
                        formatted_value = str(calculated_value)
                elif isinstance(calculated_value, (int, float)):
                    try:
                        formatted_value = format(calculated_value, cell.number_format)
                    except (ValueError, KeyError):
                        formatted_value = str(calculated_value)
                else:
                    formatted_value = str(calculated_value)
            elif isinstance(cell.value, str) and cell.value.startswith("="):
                # if cell value contains formula then that needs to be followed until a value is found
                final_value, number_format = _resolve_formula_value(
                    wb_data, cell.value, ws_data, wb
                )

                if final_value is not None:
                    if isinstance(final_value, datetime):
                        try:
                            format_str = _excel_format_to_strftime(
                                number_format or cell.number_format
                            )
                            formatted_value = final_value.strftime(format_str)
                        except (AttributeError, ValueError):
                            formatted_value = str(final_value)
                    elif isinstance(final_value, (int, float)) and number_format:
                        try:
                            formatted_value = format(final_value, number_format)
                        except (ValueError, KeyError):
                            formatted_value = str(final_value)
                    else:
                        formatted_value = str(final_value)
                else:
                    continue

            # Check against both raw and formatted values
            if formatted_value in header_positions:
                header_positions[formatted_value]["rows"].append(row_idx)
                header_positions[formatted_value]["cols"].append(col_idx)

    # Verify that headers were found and track which ones are missing
    missing_headers = []
    for header in h1_list + h2_list:
        if not header_positions[header]["rows"]:
            missing_headers.append(header)

    # Determine the intersection points
    results = CrossReferenceResult(intersections=[])
    for h1 in h1_list:
        for h2 in h2_list:
            # If either header is missing, add None as the cell reference
            if h1 in missing_headers or h2 in missing_headers:
                results.intersections.append(None)
                continue

            if min(header_positions[h1]["cols"]) < min(header_positions[h2]["cols"]):
                row = header_positions[h1]["rows"][0]  # Use the row where h1 appears
                col = header_positions[h2]["cols"][0]  # Use the column where h2 appears
            else:
                row = header_positions[h2]["rows"][0]  # Use the row where h2 appears
                col = header_positions[h1]["cols"][0]  # Use the column where h1 appears

            column_letter = openpyxl.utils.get_column_letter(col)
            cell_ref = f"{column_letter}{row}"
            results.intersections.append(cell_ref)

    return results


def _resolve_formula_value(wb_data, cell_value, current_sheet=None, data_book=None):
    """Recursively resolves formula references until finding a non-formula value.

    Args:
        wb_data: Workbook with data_only=True
        cell_value: The formula string to resolve

    Returns:
        The final resolved value and its number format
    """
    try:
        if not isinstance(cell_value, str) or not cell_value.startswith("="):
            return cell_value, None

        data_cell = None
        formula = cell_value.lstrip("=")
        if "!" in formula:
            sheet_ref, cell_ref = formula.split("!")
            sheet_name = sheet_ref.strip("'")
            ref_sheet = wb_data[sheet_name]
            ref_cell = ref_sheet[cell_ref]
            current_sheet = ref_sheet
            if data_book is not None:
                data_sheet = data_book[sheet_name]
                data_cell = data_sheet[cell_ref]
        else:
            if current_sheet is None:
                return None, None
            ref_cell = current_sheet[formula]
            if data_book is not None:
                data_sheet = data_book[current_sheet.title]
                data_cell = data_sheet[cell_ref]
        cell_to_check = ref_cell if ref_cell.value else data_cell

        # Recursively follow the reference chain
        if isinstance(cell_to_check.value, str) and cell_to_check.value.startswith("="):
            return _resolve_formula_value(
                wb_data,
                cell_to_check.value,
                cell_to_check.parent,
                data_book,
            )
        else:
            return cell_to_check.value, cell_to_check.number_format
    except Exception as e:
        print(str(e))
        return None, None


def _excel_format_to_strftime(excel_format: str) -> str:
    """Converts Excel date format to Python strftime format

    Args:
        excel_format: Excel format string (e.g., 'mmm\'yy')
    Returns:
        Python strftime format string (e.g., '%b\'%y')
    """
    # Common Excel to Python format mappings
    format_mapping = {
        "yyyy": "%Y",  # 2024
        "yy": "%y",  # 24
        "mmmm": "%B",  # January
        "mmm": "%b",  # Jan
        "mm": "%m",  # 01
        "m": "%-m",  # 1
        "dddd": "%A",  # Monday
        "ddd": "%a",  # Mon
        "dd": "%d",  # 01
        "d": "%-d",  # 1
    }

    # Handle special case for Excel's escaped characters
    python_format = excel_format.replace("\\'", "'")

    # Replace Excel format codes with Python ones
    for excel_code, python_code in format_mapping.items():
        python_format = python_format.replace(excel_code.lower(), python_code)
        python_format = python_format.replace(excel_code.upper(), python_code)

    return python_format
